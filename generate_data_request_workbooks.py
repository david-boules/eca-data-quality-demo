"""Generate one synthetic Data Request XLSX workbook per selected company."""

from __future__ import annotations

import argparse
import json
import random
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from data_request_mapping import SHEETS

ROOT = Path(__file__).resolve().parent
SOURCE_DB = ROOT / "eca_demo_large.db"
OUTPUT_DIR = ROOT / "synthetic_data_requests"
MANIFEST = "selection_manifest.json"


def select_company_ids(source_db: Path, count: int, seed: int) -> list[int]:
    with sqlite3.connect(source_db) as conn:
        ids = [r[0] for r in conn.execute("SELECT company_id FROM companies ORDER BY company_id")]
    if count > len(ids):
        raise ValueError(f"Requested {count} companies but source contains {len(ids)}")
    return sorted(random.Random(seed).sample(ids, count))


def _rows(conn: sqlite3.Connection, sql: str, params: tuple) -> list[dict]:
    cur = conn.execute(sql, params)
    return [dict(zip((d[0] for d in cur.description), row)) for row in cur]


def extract_company(conn: sqlite3.Connection, company_id: int) -> dict[str, list[dict]]:
    q: dict[str, tuple[str, tuple]] = {
        "Company Info": ("SELECT company_name,market_or_sector,establishment_year,company_type,contact_details FROM companies WHERE company_id=?", (company_id,)),
        "Financials": ("SELECT year,issued_capital,paid_in_capital,total_assets,annual_revenue,total_liabilities,total_expenses FROM company_financials WHERE company_id=? ORDER BY year", (company_id,)),
        "Activities": ("SELECT activity_name FROM company_activities WHERE company_id=? ORDER BY activity_name", (company_id,)),
        "Related Parties": ("SELECT related_party_name,contact_details FROM related_parties WHERE company_id=? ORDER BY related_party_name", (company_id,)),
        "Products": ("""SELECT p.product_name,p.brand_name,p.product_specifications,p.license_holder_company,p.manufacturer_company,p.product_source,p.product_type,p.dispensing_method,p.registration_status,p.registration_authority,p.therapeutic_purpose,p.active_ingredient,pa.alternative_product_name,pa.alternative_product_manufacturer FROM company_products cp JOIN products p ON p.product_id=cp.product_id LEFT JOIN product_alternatives pa ON pa.product_id=p.product_id WHERE cp.company_id=? ORDER BY p.product_name,pa.product_alternative_id""", (company_id,)),
        "Customers": ("SELECT customer_name,customer_type,customer_code,branch_area,governorate,relationship_start_date,relationship_end_date,phone_number FROM customers WHERE company_id=? ORDER BY customer_code", (company_id,)),
        "Imports": ("""SELECT p.product_name,su.supplier_name,i.year,i.month,i.country_of_origin,i.cif_import_price,i.quantity,i.discount_value,i.purchase_value FROM imports i JOIN company_products cp ON cp.company_product_id=i.company_product_id JOIN products p ON p.product_id=cp.product_id JOIN suppliers su ON su.supplier_id=i.supplier_id WHERE cp.company_id=? ORDER BY i.import_id""", (company_id,)),
        "Local Purchases": ("""SELECT p.product_name,su.supplier_name,x.year,x.month,x.quantity,x.discount_value,x.price_excluding_tax_and_discounts,x.purchase_value_excluding_tax_and_discounts,x.purchase_value_including_tax_and_discounts FROM purchases x JOIN company_products cp ON cp.company_product_id=x.company_product_id JOIN products p ON p.product_id=cp.product_id JOIN suppliers su ON su.supplier_id=x.supplier_id WHERE cp.company_id=? ORDER BY x.purchase_id""", (company_id,)),
        "Costs": ("""SELECT p.product_name,c.year,c.month,c.wages_and_salaries,c.financing_and_banking_costs,c.administrative_expenses,c.other_fixed_costs,c.total_fixed_cost,c.drug_purchase_cost,c.energy_cost,c.transport_cost,c.other_variable_costs,c.total_variable_cost,c.total_production_cost FROM costs c JOIN company_products cp ON cp.company_product_id=c.company_product_id JOIN products p ON p.product_id=cp.product_id WHERE cp.company_id=? ORDER BY c.cost_id""", (company_id,)),
        "Sales": ("""SELECT p.product_name,cu.customer_code,s.year,s.month,s.sales_quantity,s.returned_quantity,s.sale_price_excluding_tax_and_discounts,s.customer_discount_value,s.sales_value_excluding_tax_and_discounts,s.sales_value_including_tax_and_discounts FROM sales s JOIN company_products cp ON cp.company_product_id=s.company_product_id JOIN products p ON p.product_id=cp.product_id JOIN customers cu ON cu.customer_id=s.customer_id WHERE cp.company_id=? ORDER BY s.sale_id""", (company_id,)),
        "Tenders": ("""SELECT 'T'||printf('%08d',t.tender_id) tender_key,t.contractual_operation_type,t.tendering_entity_name,t.tender_date,p.product_name,ti.price_excluding_tax,ti.sales_quantity,ti.sales_value_excluding_tax,ti.sales_value_including_tax FROM tenders t JOIN tender_items ti ON ti.tender_id=t.tender_id JOIN company_products cp ON cp.company_product_id=ti.company_product_id JOIN products p ON p.product_id=cp.product_id WHERE t.company_id=? ORDER BY t.tender_id,ti.tender_item_id""", (company_id,)),
        "Exports": ("""SELECT p.product_name,e.year,e.month,e.recipient_company_name,e.destination_country,e.export_price_excluding_tax,e.export_quantity,e.export_value_excluding_tax,e.export_value_including_tax FROM exports e JOIN company_products cp ON cp.company_product_id=e.company_product_id JOIN products p ON p.product_id=cp.product_id WHERE cp.company_id=? ORDER BY e.export_id""", (company_id,)),
        "Storage Capacity": ("""SELECT w.warehouse_name,w.warehouse_area,sc.year,sc.storage_capacity FROM warehouses w JOIN storage_capacity sc ON sc.warehouse_id=w.warehouse_id WHERE w.company_id=? ORDER BY w.warehouse_name,sc.year""", (company_id,)),
        "Inventory": ("""SELECT w.warehouse_name,p.product_name,i.year,i.inventory_quantity,i.inventory_value_excluding_tax,i.inventory_value_including_tax FROM inventory i JOIN warehouses w ON w.warehouse_id=i.warehouse_id JOIN company_products cp ON cp.company_product_id=i.company_product_id JOIN products p ON p.product_id=cp.product_id WHERE w.company_id=? ORDER BY i.inventory_id""", (company_id,)),
    }
    return {sheet: _rows(conn, sql, params) for sheet, (sql, params) in q.items()}


def write_workbook(path: Path, data: dict[str, list[dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, mapping in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.append(list(mapping.columns))
        for row in data[sheet_name]:
            ws.append([row.get(c) for c in mapping.columns])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="17365D")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            width = min(32, max(12, max(len(str(c.value or "")) for c in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
    wb.properties.title = "Synthetic ECA Data Request Demonstration Input"
    wb.properties.subject = "Fully synthetic data; not an internal or confidential template"
    wb.save(path)


def generate_workbooks(source_db: Path, output_dir: Path, companies: int, seed: int, overwrite: bool = False) -> dict:
    if not source_db.exists():
        raise FileNotFoundError(f"Source database not found: {source_db}")
    output_dir.mkdir(parents=True, exist_ok=True)
    existing = list(output_dir.glob("company_*.xlsx"))
    if existing and not overwrite:
        raise FileExistsError(f"{output_dir} already contains workbooks; use --overwrite")
    for path in existing:
        path.unlink()
    ids = select_company_ids(source_db, companies, seed)
    selected = []
    with sqlite3.connect(source_db) as conn:
        for index, company_id in enumerate(ids, 1):
            data = extract_company(conn, company_id)
            filename = f"company_{index:04d}.xlsx"
            write_workbook(output_dir / filename, data)
            selected.append({"file": filename, "source_company_id": company_id,
                             "company_name": data["Company Info"][0]["company_name"]})
    manifest = {"synthetic": True, "source_database": str(source_db.resolve()), "seed": seed,
                "company_count": companies, "companies": selected}
    (output_dir / MANIFEST).write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return manifest


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--source-db", type=Path, default=SOURCE_DB)
    p.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    p.add_argument("--companies", type=int, default=100)
    p.add_argument("--seed", type=int, default=42)
    p.add_argument("--overwrite", action="store_true")
    args = p.parse_args()
    result = generate_workbooks(args.source_db, args.output_dir, args.companies, args.seed, args.overwrite)
    print(f"Generated {result['company_count']} synthetic workbooks in {args.output_dir}")


if __name__ == "__main__":
    main()
