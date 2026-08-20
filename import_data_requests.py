"""Validate and transactionally import synthetic Data Request workbooks."""

from __future__ import annotations

import argparse
import json
import math
import sqlite3
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from create_database import SCHEMA_SQL
from data_request_mapping import INTEGER_FIELDS, NONNEGATIVE_FIELDS, REQUIRED_SHEETS, SHEETS

ROOT = Path(__file__).resolve().parent


@dataclass
class Issue:
    severity: str
    sheet: str
    row: int | None
    field: str | None
    message: str


@dataclass
class ImportResult:
    file: str
    status: str
    company_name: str | None = None
    rows_inserted: dict[str, int] = field(default_factory=dict)
    rows_rejected: int = 0
    issues: list[Issue] = field(default_factory=list)

    def to_dict(self) -> dict:
        value = asdict(self)
        value["validation_failures"] = sum(i.severity in {"fatal", "error"} for i in self.issues)
        value["warnings"] = sum(i.severity == "warning" for i in self.issues)
        value["rows_inserted_total"] = sum(self.rows_inserted.values())
        return value


def _clean(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def read_workbook(path: Path) -> tuple[dict[str, list[dict]], list[Issue]]:
    issues: list[Issue] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {}, [Issue("fatal", "workbook", None, None, f"Malformed or unreadable workbook: {exc}")]
    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    for sheet in missing:
        issues.append(Issue("fatal", sheet, None, None, "Required sheet is missing"))
    data: dict[str, list[dict]] = {}
    for sheet_name, mapping in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(v).strip() if v is not None else "" for v in next(rows)]
        except StopIteration:
            issues.append(Issue("fatal", sheet_name, 1, None, "Sheet is empty"))
            data[sheet_name] = []
            continue
        missing_cols = [c for c in mapping.columns if c not in headers]
        for col in missing_cols:
            issues.append(Issue("fatal", sheet_name, 1, col, "Required mapped column is missing"))
        parsed = []
        for row_number, values in enumerate(rows, 2):
            if all(v is None or str(v).strip() == "" for v in values):
                continue
            record = {h: _clean(values[i]) if i < len(values) else None for i, h in enumerate(headers) if h}
            parsed.append(record)
            for required in mapping.required:
                if record.get(required) is None:
                    issues.append(Issue("error", sheet_name, row_number, required, "Required value is missing"))
            for name, value in record.items():
                if value is None:
                    continue
                if name in INTEGER_FIELDS:
                    if isinstance(value, bool) or not isinstance(value, (int, float)) or not float(value).is_integer():
                        issues.append(Issue("error", sheet_name, row_number, name, "Expected an integer"))
                    else:
                        record[name] = int(value)
                if name in NONNEGATIVE_FIELDS:
                    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
                        issues.append(Issue("error", sheet_name, row_number, name, "Expected a numeric value"))
                    elif value < 0:
                        issues.append(Issue("error", sheet_name, row_number, name, "Negative values are not allowed"))
            if isinstance(record.get("month"), int) and not 1 <= record["month"] <= 12:
                issues.append(Issue("error", sheet_name, row_number, "month", "Month must be between 1 and 12"))
            if sheet_name == "Sales" and isinstance(record.get("returned_quantity"), (int, float)) and isinstance(record.get("sales_quantity"), (int, float)) and record["returned_quantity"] > record["sales_quantity"]:
                issues.append(Issue("error", sheet_name, row_number, "returned_quantity", "Returns exceed sales quantity"))
        data[sheet_name] = parsed
    wb.close()
    if len(data.get("Company Info", [])) != 1:
        issues.append(Issue("fatal", "Company Info", None, None, "Exactly one company row is required"))
    # Exact duplicate workbook rows are almost always accidental; product rows can repeat for alternatives.
    for sheet, records in data.items():
        if sheet == "Products":
            continue
        seen: set[tuple] = set()
        for idx, record in enumerate(records, 2):
            key = tuple((k, record.get(k)) for k in SHEETS[sheet].columns)
            if key in seen:
                issues.append(Issue("error", sheet, idx, None, "Duplicate row"))
            seen.add(key)

    tender_metadata: dict[str, tuple[tuple[Any, ...], int]] = {}
    tender_fields = ("contractual_operation_type", "tendering_entity_name", "tender_date")
    for row_number, record in enumerate(data.get("Tenders", []), 2):
        key = record.get("tender_key")
        if key is None:
            continue
        metadata = tuple(record.get(field) for field in tender_fields)
        previous = tender_metadata.get(str(key))
        if previous and previous[0] != metadata:
            differences = [field for field, old, new in zip(tender_fields, previous[0], metadata) if old != new]
            issues.append(Issue(
                "error", "Tenders", row_number, "tender_key",
                f"Tender {key!r} conflicts with row {previous[1]} for: {', '.join(differences)}",
            ))
        else:
            tender_metadata[str(key)] = (metadata, row_number)
    return data, issues


def _insert(conn: sqlite3.Connection, table: str, values: dict, counts: dict[str, int]) -> int:
    cols = list(values)
    cur = conn.execute(f"INSERT INTO {table} ({','.join(cols)}) VALUES ({','.join('?' for _ in cols)})", tuple(values[c] for c in cols))
    counts[table] = counts.get(table, 0) + 1
    return int(cur.lastrowid)


def _resolve_product(conn, record, counts) -> int:
    base_cols = SHEETS["Products"].columns[:12]
    row = conn.execute(
        f"SELECT product_id,{','.join(base_cols)} FROM products WHERE product_name=? ORDER BY product_id LIMIT 1",
        (record["product_name"],),
    ).fetchone()
    if row:
        existing = dict(zip(("product_id", *base_cols), row))
        conflicts = [
            column for column in base_cols if column != "product_name"
            and record.get(column) is not None and existing.get(column) is not None
            and record[column] != existing[column]
        ]
        if conflicts:
            details = ", ".join(
                f"{column} (existing={existing[column]!r}, incoming={record[column]!r})"
                for column in conflicts
            )
            raise ValueError(
                f"Product metadata conflict for exact product_name {record['product_name']!r}: {details}"
            )
        return row[0]
    return _insert(conn, "products", {c: record.get(c) for c in base_cols}, counts)


def _resolve_supplier(conn, name, counts) -> int:
    row = conn.execute("SELECT supplier_id FROM suppliers WHERE supplier_name=?", (name,)).fetchone()
    return row[0] if row else _insert(conn, "suppliers", {"supplier_name": name}, counts)


def import_workbook(path: Path, db_path: Path) -> ImportResult:
    data, issues = read_workbook(path)
    company_name = data.get("Company Info", [{}])[0].get("company_name") if data.get("Company Info") else None
    result = ImportResult(str(path), "FAILED", company_name, issues=issues)
    if any(i.severity in {"fatal", "error"} for i in issues):
        result.rows_rejected = sum(len(v) for v in data.values())
        return result
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys=ON")
    counts: dict[str, int] = {}
    try:
        conn.execute("BEGIN")
        if conn.execute("SELECT 1 FROM companies WHERE company_name=?", (company_name,)).fetchone():
            raise ValueError("This company workbook has already been imported; duplicate rejected")
        company_id = _insert(conn, "companies", data["Company Info"][0], counts)
        for row in data["Financials"]:
            _insert(conn, "company_financials", {"company_id": company_id, **row}, counts)
        for row in data["Activities"]:
            _insert(conn, "company_activities", {"company_id": company_id, **row}, counts)
        for row in data["Related Parties"]:
            _insert(conn, "related_parties", {"company_id": company_id, **row}, counts)
        product_ids: dict[str, int] = {}
        cp_ids: dict[str, int] = {}
        for row in data["Products"]:
            name = row["product_name"]
            if name not in product_ids:
                product_ids[name] = _resolve_product(conn, row, counts)
                cp_ids[name] = _insert(conn, "company_products", {"company_id": company_id, "product_id": product_ids[name]}, counts)
            if row.get("alternative_product_name") and not conn.execute("SELECT 1 FROM product_alternatives WHERE product_id=? AND alternative_product_name=?", (product_ids[name], row["alternative_product_name"])).fetchone():
                _insert(conn, "product_alternatives", {"product_id": product_ids[name], "alternative_product_name": row["alternative_product_name"], "alternative_product_manufacturer": row.get("alternative_product_manufacturer")}, counts)
        customer_ids = {}
        for row in data["Customers"]:
            customer_ids[row["customer_code"]] = _insert(conn, "customers", {"company_id": company_id, **row}, counts)
        warehouses = {}
        for row in data["Storage Capacity"]:
            name = row["warehouse_name"]
            if name not in warehouses:
                warehouses[name] = _insert(conn, "warehouses", {"company_id": company_id, "warehouse_name": name, "warehouse_area": row.get("warehouse_area")}, counts)
            _insert(conn, "storage_capacity", {"warehouse_id": warehouses[name], "year": row["year"], "storage_capacity": row["storage_capacity"]}, counts)
        def cp(name: str) -> int:
            if name not in cp_ids:
                raise ValueError(f"Unknown product reference for this company: {name}")
            return cp_ids[name]
        for row in data["Imports"]:
            supplier = _resolve_supplier(conn, row.pop("supplier_name"), counts)
            _insert(conn, "imports", {"company_product_id": cp(row.pop("product_name")), "supplier_id": supplier, **row}, counts)
        for row in data["Local Purchases"]:
            supplier = _resolve_supplier(conn, row.pop("supplier_name"), counts)
            _insert(conn, "purchases", {"company_product_id": cp(row.pop("product_name")), "supplier_id": supplier, **row}, counts)
        for sheet, table in (("Costs", "costs"), ("Exports", "exports")):
            for row in data[sheet]:
                _insert(conn, table, {"company_product_id": cp(row.pop("product_name")), **row}, counts)
        for row in data["Sales"]:
            product = row.pop("product_name"); customer = row.pop("customer_code")
            if customer not in customer_ids:
                raise ValueError(f"Unknown customer_code reference: {customer}")
            _insert(conn, "sales", {"company_product_id": cp(product), "customer_id": customer_ids[customer], **row}, counts)
        tender_ids = {}
        for row in data["Tenders"]:
            key = row.pop("tender_key"); product = row.pop("product_name")
            if key not in tender_ids:
                tender_ids[key] = _insert(conn, "tenders", {"company_id": company_id, "contractual_operation_type": row.pop("contractual_operation_type"), "tendering_entity_name": row.pop("tendering_entity_name"), "tender_date": row.pop("tender_date")}, counts)
            else:
                for col in ("contractual_operation_type", "tendering_entity_name", "tender_date"):
                    row.pop(col)
            _insert(conn, "tender_items", {"tender_id": tender_ids[key], "company_product_id": cp(product), **row}, counts)
        for row in data["Inventory"]:
            warehouse = row.pop("warehouse_name"); product = row.pop("product_name")
            if warehouse not in warehouses:
                raise ValueError(f"Unknown warehouse reference: {warehouse}")
            _insert(conn, "inventory", {"warehouse_id": warehouses[warehouse], "company_product_id": cp(product), **row}, counts)
        fk = conn.execute("PRAGMA foreign_key_check").fetchall()
        if fk:
            raise ValueError(f"Foreign-key verification failed: {fk[:3]}")
        conn.commit()
        result.status = "SUCCESS"; result.rows_inserted = counts
    except (sqlite3.Error, ValueError) as exc:
        conn.rollback()
        result.issues.append(Issue("error", "import", None, None, str(exc)))
        result.rows_rejected = sum(len(v) for v in data.values())
    finally:
        conn.close()
    return result


def batch_import(directory: Path, db_path: Path, report_path: Path | None = None) -> dict:
    results = [import_workbook(path, db_path) for path in sorted(directory.glob("*.xlsx"))]
    report = {"files_processed": len(results), "successful_imports": sum(r.status == "SUCCESS" for r in results),
              "failed_imports": sum(r.status != "SUCCESS" for r in results),
              "warnings": sum(sum(i.severity == "warning" for i in r.issues) for r in results),
              "validation_failures": sum(sum(i.severity in {"fatal", "error"} for i in r.issues) for r in results),
              "rows_inserted": sum(sum(r.rows_inserted.values()) for r in results),
              "rows_rejected": sum(r.rows_rejected for r in results), "results": [r.to_dict() for r in results]}
    if report_path:
        report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    return report


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--file", type=Path); g.add_argument("--directory", type=Path)
    p.add_argument("--db", type=Path, default=ROOT / "eca_demo.db")
    p.add_argument("--report", type=Path, default=ROOT / "import_report.json")
    p.add_argument("--create-fresh", action="store_true")
    args = p.parse_args()
    if args.create_fresh:
        if args.db.exists(): args.db.unlink()
        with sqlite3.connect(args.db) as conn: conn.executescript(SCHEMA_SQL)
    if not args.db.exists():
        raise FileNotFoundError("Target database does not exist; use --create-fresh or create_database.py")
    if args.file:
        report = import_workbook(args.file, args.db).to_dict()
        print(json.dumps(report, indent=2))
    else:
        report = batch_import(args.directory, args.db, args.report)
        for item in report["results"]:
            detail = "" if item["status"] == "SUCCESS" else f" — {item['validation_failures']} validation error(s)"
            print(f"{Path(item['file']).name:<24} {item['status']}{detail}")
        print(json.dumps({key: value for key, value in report.items() if key != "results"}, indent=2))


if __name__ == "__main__":
    main()
