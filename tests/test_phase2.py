from __future__ import annotations
import json
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

from data_request_mapping import REQUIRED_SHEETS
from db import Database
from generate_data_request_workbooks import generate_workbooks, select_company_ids
from import_data_requests import batch_import, import_workbook
import launch_app
from round_trip_validation import validate_round_trip
from validate_database import EXPECTED_TABLES, get_duplicate_financial_pair


def make_batch(source_db, tmp_path, count=3):
    out = tmp_path / "books"
    manifest = generate_workbooks(source_db, out, count, 42)
    return out, manifest


def test_database_creation_and_foreign_keys(target_db):
    with sqlite3.connect(target_db) as conn:
        assert conn.execute("PRAGMA integrity_check").fetchone()[0] == "ok"
        assert conn.execute("PRAGMA foreign_key_check").fetchall() == []
        assert conn.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'").fetchone()[0] == 19
    assert Database(target_db).preflight(EXPECTED_TABLES) == (True,"ok")


def test_database_preflight_handles_missing_and_incomplete_files(tmp_path):
    ok,message=Database(tmp_path/"missing.db").preflight(EXPECTED_TABLES)
    assert not ok and "not found" in message
    incomplete=tmp_path/"incomplete.db"
    with sqlite3.connect(incomplete) as conn: conn.execute("CREATE TABLE placeholder(id INTEGER)")
    ok,message=Database(incomplete).preflight(EXPECTED_TABLES)
    assert not ok and "Missing tables" in message


def test_launcher_selects_next_available_port(monkeypatch):
    monkeypatch.setattr(launch_app,"port_is_available",lambda port: port==8503)
    assert launch_app.choose_port(8501)==8503


def test_deterministic_selection(source_db):
    assert select_company_ids(source_db, 5, 42) == select_company_ids(source_db, 5, 42)
    assert select_company_ids(source_db, 5, 42) != select_company_ids(source_db, 5, 43)


def test_generator_count_sheets_isolation_and_source_values(source_db, tmp_path):
    out, manifest = make_batch(source_db, tmp_path, 3)
    files = sorted(out.glob("*.xlsx")); assert len(files) == 3
    names=[]
    for file, selected in zip(files, manifest["companies"]):
        wb=load_workbook(file,read_only=True,data_only=True)
        assert set(REQUIRED_SHEETS).issubset(wb.sheetnames)
        name=wb["Company Info"]["A2"].value; names.append(name); assert name == selected["company_name"]
        workbook_products={r[0] for r in wb["Products"].iter_rows(min_row=2,values_only=True)}
        with sqlite3.connect(source_db) as conn:
            source_products={r[0] for r in conn.execute("SELECT p.product_name FROM products p JOIN company_products cp ON cp.product_id=p.product_id JOIN companies c ON c.company_id=cp.company_id WHERE c.company_name=?",(name,))}
        assert workbook_products == source_products
    assert len(set(names)) == 3


def test_valid_single_multiple_duplicate_and_roundtrip(source_db, target_db, tmp_path):
    out, _ = make_batch(source_db, tmp_path, 3)
    first=sorted(out.glob("*.xlsx"))[0]
    assert import_workbook(first,target_db).status == "SUCCESS"
    duplicate=import_workbook(first,target_db); assert duplicate.status == "FAILED"; assert "duplicate rejected" in duplicate.issues[-1].message
    report=batch_import(out,target_db,tmp_path/"report.json")
    assert report["successful_imports"] == 2 and report["failed_imports"] == 1
    rt=validate_round_trip(source_db,target_db,out/"selection_manifest.json")
    assert rt["passed"], rt["mismatches"]
    assert rt["representative_sales_query"]["year"] == 2026
    assert rt["representative_field_checks"]["passed"]
    assert rt["representative_field_checks"]["performed"] > 0


def test_missing_sheet_and_rollback(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); book=next(out.glob("*.xlsx")); wb=load_workbook(book); del wb["Sales"]; broken=tmp_path/"missing.xlsx"; wb.save(broken)
    result=import_workbook(broken,target_db); assert result.status=="FAILED"; assert any("Required sheet" in i.message for i in result.issues)
    with sqlite3.connect(target_db) as conn: assert conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]==0


def test_missing_value_invalid_numeric_and_relationship(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); original=next(out.glob("*.xlsx"))
    for kind, sheet, cell, value, expected in [
        ("missing","Company Info","A2",None,"Required value"),
        ("numeric","Sales","C2","not-a-year","Expected an integer"),
        ("relationship","Sales","A2","Unknown Product","Unknown product reference")]:
        wb=load_workbook(original); wb[sheet][cell]=value; path=tmp_path/f"{kind}.xlsx"; wb.save(path)
        result=import_workbook(path,target_db); assert result.status=="FAILED"; assert any(expected in i.message for i in result.issues)
        with sqlite3.connect(target_db) as conn: assert conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]==0


def test_malformed_workbook(target_db,tmp_path):
    path=tmp_path/"bad.xlsx"; path.write_text("not xlsx")
    result=import_workbook(path,target_db); assert result.status=="FAILED"; assert "Malformed" in result.issues[0].message


def test_mapping_representative_fields(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); assert import_workbook(next(out.glob("*.xlsx")),target_db).status=="SUCCESS"
    with sqlite3.connect(target_db) as conn:
        assert conn.execute("SELECT COUNT(*) FROM company_financials").fetchone()[0] > 0
        assert conn.execute("SELECT COUNT(*) FROM company_products").fetchone()[0] > 0
        assert conn.execute("SELECT COUNT(*) FROM sales").fetchone()[0] > 0
        assert conn.execute("PRAGMA foreign_key_check").fetchall()==[]


def test_shared_product_metadata_conflict_is_rejected(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); book=next(out.glob("*.xlsx"))
    wb=load_workbook(book,read_only=True,data_only=True)
    product_name=wb["Products"]["A2"].value; incoming_brand=wb["Products"]["B2"].value; wb.close()
    conflict_brand = f"{incoming_brand} CONFLICT"
    with sqlite3.connect(target_db) as conn:
        conn.execute("INSERT INTO products(product_name,brand_name) VALUES(?,?)",(product_name,conflict_brand))
    result=import_workbook(book,target_db)
    assert result.status=="FAILED"
    assert any("Product metadata conflict" in issue.message and "brand_name" in issue.message for issue in result.issues)
    with sqlite3.connect(target_db) as conn: assert conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]==0


def test_repeated_tender_metadata_conflict_is_rejected(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); book=next(out.glob("*.xlsx")); wb=load_workbook(book)
    product_name=wb["Products"]["A2"].value
    ws=wb["Tenders"]
    ws.append(["TEST-TENDER","Tender","Entity A","2026-01-10",product_name,10,2,20,22])
    ws.append(["TEST-TENDER","Tender","Entity B","2026-01-10",product_name,10,3,30,33])
    path=tmp_path/"tender_conflict.xlsx"; wb.save(path)
    result=import_workbook(path,target_db)
    assert result.status=="FAILED"
    assert any("Tender 'TEST-TENDER' conflicts" in issue.message and "tendering_entity_name" in issue.message for issue in result.issues)
    with sqlite3.connect(target_db) as conn: assert conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]==0


def test_dynamic_duplicate_financial_validation(source_db,tmp_path):
    isolated=tmp_path/"financial_constraint.db"
    with sqlite3.connect(source_db) as source, sqlite3.connect(isolated) as target: source.backup(target)
    with sqlite3.connect(isolated) as conn:
        conn.execute("DELETE FROM company_financials WHERE year=2025")
        pair=get_duplicate_financial_pair(conn)
        assert pair[1] != 2025
        try: conn.execute("INSERT INTO company_financials(company_id,year,annual_revenue) VALUES(?,?,?)",(*pair,1))
        except sqlite3.IntegrityError: pass
        else: raise AssertionError("dynamic existing pair was not rejected")


def test_representative_field_roundtrip_detects_difference(source_db,target_db,tmp_path):
    out,_=make_batch(source_db,tmp_path,1); assert import_workbook(next(out.glob("*.xlsx")),target_db).status=="SUCCESS"
    good=validate_round_trip(source_db,target_db,out/"selection_manifest.json")
    assert good["representative_field_checks"]["passed"]
    with sqlite3.connect(target_db) as conn: conn.execute("UPDATE companies SET company_type='Changed Type'")
    bad=validate_round_trip(source_db,target_db,out/"selection_manifest.json")
    assert not bad["passed"]
    assert any(item["metric"]=="representative_field:company" for item in bad["mismatches"])


def test_streamlit_pages_have_no_runtime_exceptions(source_db,target_db,tmp_path,monkeypatch):
    from streamlit.testing.v1 import AppTest
    out,_=make_batch(source_db,tmp_path,1); assert import_workbook(next(out.glob("*.xlsx")),target_db).status=="SUCCESS"
    monkeypatch.setenv("ECA_DB_PATH",str(target_db))
    app = AppTest.from_file(str(Path(__file__).resolve().parents[1] / "app.py"), default_timeout=20).run()
    assert not app.exception
    for page in ["Company Explorer", "Product Explorer", "Data Explorer", "Sales / Market Analysis", "Data Quality / Import Status", "Import Data", "System Guide"]:
        app.sidebar.radio[0].set_value(page); app.run(); assert not app.exception, page


def test_streamlit_empty_database_and_filter_capabilities(target_db,monkeypatch):
    from streamlit.testing.v1 import AppTest
    monkeypatch.setenv("ECA_DB_PATH",str(target_db))
    app=AppTest.from_file(str(Path(__file__).resolve().parents[1]/"app.py"),default_timeout=20).run()
    for page in ["Overview","Company Explorer","Product Explorer","Data Explorer","Sales / Market Analysis","Data Quality / Import Status","Import Data","System Guide"]:
        app.sidebar.radio[0].set_value(page); app.run(); assert not app.exception,page
    app.sidebar.radio[0].set_value("Company Explorer"); app.run(); assert any("No companies" in item.value for item in app.info)
    app.sidebar.radio[0].set_value("Product Explorer"); app.run(); assert any("No products" in item.value for item in app.info)
    app.sidebar.radio[0].set_value("Data Explorer"); app.run()
    assert [item.label for item in app.selectbox]==["Dataset"]
    app.selectbox[0].set_value("sales"); app.run()
    assert [item.label for item in app.selectbox]==["Dataset","Company","Product","Year"]
